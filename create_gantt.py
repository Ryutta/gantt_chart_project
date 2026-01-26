import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# 1. Prepare Data
data = [
    {"Task": "Requirement Analysis", "Start": "2023-10-01", "End": "2023-10-05"},
    {"Task": "Design", "Start": "2023-10-06", "End": "2023-10-10"},
    {"Task": "Implementation", "Start": "2023-10-11", "End": "2023-10-20"},
    {"Task": "Testing", "Start": "2023-10-21", "End": "2023-10-25"},
    {"Task": "Deployment", "Start": "2023-10-26", "End": "2023-10-28"},
]

df = pd.DataFrame(data)
df['Start'] = pd.to_datetime(df['Start'])
df['End'] = pd.to_datetime(df['End'])

# Calculate the overall start and end date for the calendar view
min_date = df['Start'].min()
max_date = df['End'].max()
date_range = pd.date_range(start=min_date, end=max_date)

# 2. Write to Excel
file_name = 'gantt.xlsx'
# We write the task data first
with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Gantt', index=False)

# 3. Open with openpyxl for formatting
wb = load_workbook(file_name)
ws = wb['Gantt']

# Add date headers starting from column D (index 4)
# Columns: A=Task, B=Start, C=End. So dates start at D.
# However, to facilitate conditional formatting, let's just use columns D onwards.

# Write date headers
start_col_idx = 4 # D column (1-based index)
for i, single_date in enumerate(date_range):
    col_idx = start_col_idx + i
    cell = ws.cell(row=1, column=col_idx)
    cell.value = single_date
    cell.number_format = 'd' # Show only day
    # Adjust column width
    ws.column_dimensions[get_column_letter(col_idx)].width = 3

# Add Month/Year info potentially in a row above or just keep it simple.
# For simplicity, we just put the date in row 1.

# 4. Apply Styles and Conditional Formatting
blue_fill = PatternFill(start_color='00CCFF', end_color='00CCFF', fill_type='solid')
header_font = Font(bold=True)
center_alignment = Alignment(horizontal='center')

# Style headers
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = center_alignment

# Adjust Task column width
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12

# Conditional Formatting
# Formula: =AND(D$1>=$B2, D$1<=$C2)
# Note: In openpyxl, rules are applied to a range.
# We need to apply this to the grid area: D2 : [EndColumn][EndRow]

last_row = len(data) + 1
last_col_letter = get_column_letter(start_col_idx + len(date_range) - 1)
start_cell_addr = f"D2"
end_cell_addr = f"{last_col_letter}{last_row}"
range_string = f"{start_cell_addr}:{end_cell_addr}"

# The formula needs to be relative to the top-left cell of the range (D2).
# D$1 refers to the date in the header row.
# $B2 refers to the Start Date in the current row.
# $C2 refers to the End Date in the current row.
formula = f"AND(D$1>=$B2, D$1<=$C2)"

rule = FormulaRule(formula=[formula], stopIfTrue=True, fill=blue_fill)
ws.conditional_formatting.add(range_string, rule)

# Also format the date columns B and C to show readable dates
for row in range(2, last_row + 1):
    ws[f'B{row}'].number_format = 'yyyy-mm-dd'
    ws[f'C{row}'].number_format = 'yyyy-mm-dd'

wb.save(file_name)
print(f"Gantt chart saved to {file_name}")
